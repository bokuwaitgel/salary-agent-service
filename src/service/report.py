"""Generate Excel reports from salary analysis tables."""
import os
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional

from dotenv import load_dotenv
from sqlalchemy import Column, DateTime, Float, Integer, String, Text, create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

load_dotenv()

Base = declarative_base()

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BLUE = "1F4E79"
MID_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "548235"
ACCENT_ORANGE = "ED7D31"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
MEDIUM_GREY = "D9D9D9"

# ── Reusable styles ────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, color=MID_BLUE, size=11)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
SALARY_FORMAT = '#,##0₮'
NUMBER_FORMAT = '#,##0'

THIN_BORDER = Border(
    left=Side(style="thin", color=MEDIUM_GREY),
    right=Side(style="thin", color=MEDIUM_GREY),
    top=Side(style="thin", color=MEDIUM_GREY),
    bottom=Side(style="thin", color=MEDIUM_GREY),
)

ROW_FILL_EVEN = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# Column-width presets per type
WIDTH_PRESETS: Dict[str, float] = {
    "name": 32,       # job category / level names (Mongolian text)
    "salary": 18,     # salary numbers
    "count": 12,      # small integers
    "period": 10,     # month / year
    "text_short": 28, # requirement / bonus names
    "text_long": 55,  # details / descriptions
}


class JobCategorySalary(Base):
    __tablename__ = "job_category_salary"
    id = Column(Integer, primary_key=True, autoincrement=True)
    job_category = Column(String(255), nullable=False)
    min_salary = Column(Float)
    max_salary = Column(Float)
    average_salary = Column(Float)
    requirements_details = Column(Text)
    bonus_details = Column(Text)
    job_count = Column(Integer)
    source_zangia = Column(Integer)
    source_lambda = Column(Integer)
    month = Column(Integer)
    year = Column(Integer)
    created_at = Column(DateTime)


class JobCategoryRequirement(Base):
    __tablename__ = "job_category_requirement"
    id = Column(Integer, primary_key=True, autoincrement=True)
    job_category = Column(String(255), nullable=False)
    requirement_name = Column(String(255))
    requirement_details = Column(Text)
    created_at = Column(DateTime)


class JobCategoryBonus(Base):
    __tablename__ = "job_category_bonus"
    id = Column(Integer, primary_key=True, autoincrement=True)
    job_category = Column(String(255), nullable=False)
    bonus_name = Column(String(255))
    bonus_description = Column(Text)
    created_at = Column(DateTime)


class JobLevelSalary(Base):
    __tablename__ = "job_level_salary"
    id = Column(Integer, primary_key=True, autoincrement=True)
    job_level = Column(String(255), nullable=False)
    min_salary = Column(Float)
    max_salary = Column(Float)
    average_salary = Column(Float)
    requirements_details = Column(Text)
    bonus_details = Column(Text)
    job_count = Column(Integer)
    source_zangia = Column(Integer)
    source_lambda = Column(Integer)
    month = Column(Integer)
    year = Column(Integer)
    created_at = Column(DateTime)


class JobLevelRequirement(Base):
    __tablename__ = "job_level_requirement"
    id = Column(Integer, primary_key=True, autoincrement=True)
    job_level = Column(String(255), nullable=False)
    requirement_name = Column(String(255))
    requirement_details = Column(Text)
    created_at = Column(DateTime)


class JobLevelBonus(Base):
    __tablename__ = "job_level_bonus"
    id = Column(Integer, primary_key=True, autoincrement=True)
    job_level = Column(String(255), nullable=False)
    bonus_name = Column(String(255))
    bonus_description = Column(Text)
    created_at = Column(DateTime)


# ── Helpers ─────────────────────────────────────────────────────────────────

def _get_engine():
    conn_str = os.getenv("DATABASE_URI", "sqlite:///products.db")
    return create_engine(conn_str, pool_pre_ping=True)


def _apply_header_style(ws, col_count: int) -> None:
    """Style the first row as a professional header band."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30


def _apply_body_style(ws, row_count: int, col_count: int,
                       salary_cols: Optional[List[int]] = None,
                       number_cols: Optional[List[int]] = None) -> None:
    """Style data rows with alternating fills, borders, number formats."""
    salary_cols = salary_cols or []
    number_cols = number_cols or []

    for row_idx in range(2, row_count + 2):  # data starts at row 2
        fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        # Auto row height: estimate based on longest cell text
        max_lines = 1
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = BODY_ALIGNMENT

            if col_idx in salary_cols and isinstance(cell.value, (int, float)):
                cell.number_format = SALARY_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif col_idx in number_cols and isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Estimate wrapped line count for row height
            if cell.value:
                col_width = ws.column_dimensions[get_column_letter(col_idx)].width or 12
                text_len = len(str(cell.value))
                lines = max(1, int(text_len / (col_width * 1.1)) + 1)
                if lines > max_lines:
                    max_lines = lines

        ws.row_dimensions[row_idx].height = max(15, min(max_lines * 15, 120))


def _set_column_widths(ws, widths: List[float]) -> None:
    """Set explicit column widths."""
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_sheet(ws, title: str, headers: List[str],
                 rows: Iterable[List[Any]],
                 col_widths: Optional[List[float]] = None,
                 salary_cols: Optional[List[int]] = None,
                 number_cols: Optional[List[int]] = None) -> int:
    """Write a professional data sheet. Returns the number of data rows."""
    ws.title = title

    # Write headers
    ws.append(headers)
    col_count = len(headers)

    # Write data
    data_rows = list(rows)
    for row in data_rows:
        ws.append(list(row))

    # Apply widths
    if col_widths:
        _set_column_widths(ws, col_widths)
    else:
        # Fallback auto-width with generous padding
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(
                (len(str(c.value)) if c.value else 0 for c in col_cells), default=8
            )
            ws.column_dimensions[col_letter].width = min(max_len + 6, 55)

    # Styles
    _apply_header_style(ws, col_count)
    _apply_body_style(ws, len(data_rows), col_count,
                      salary_cols=salary_cols, number_cols=number_cols)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Sheet tab colour
    ws.sheet_properties.tabColor = MID_BLUE

    return len(data_rows)


def _add_bar_chart(ws, title: str, data_row_count: int,
                   data_min_col: int, data_max_col: int,
                   cat_col: int = 1, anchor: str = "L2") -> None:
    """Add a styled bar chart to the sheet, well away from data columns."""
    if data_row_count <= 0:
        return

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Цалин (₮)"
    chart.y_axis.numFmt = '#,##0'
    chart.x_axis.title = None
    chart.x_axis.tickLblPos = "low"

    data = Reference(ws, min_col=data_min_col, max_col=data_max_col,
                     min_row=1, max_row=data_row_count + 1)
    categories = Reference(ws, min_col=cat_col, min_row=2,
                           max_row=data_row_count + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Colours for series
    colours = [MID_BLUE, ACCENT_GREEN, ACCENT_ORANGE]
    for i, s in enumerate(chart.series):
        s.graphicalProperties.solidFill = colours[i % len(colours)]

    chart.height = 12  # cm
    chart.width = 24   # cm
    chart.legend.position = "b"

    ws.add_chart(chart, anchor)


def _add_pie_chart(ws, title: str, data_row_count: int,
                   data_col: int, cat_col: int = 1,
                   anchor: str = "L20") -> None:
    """Add a styled pie chart below the bar chart."""
    if data_row_count <= 0:
        return

    chart = PieChart()
    chart.style = 10
    chart.title = title

    data = Reference(ws, min_col=data_col, min_row=1,
                     max_row=data_row_count + 1)
    categories = Reference(ws, min_col=cat_col, min_row=2,
                           max_row=data_row_count + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showVal = False

    chart.height = 12
    chart.width = 18
    if chart.legend is not None:
        chart.legend.position = "r"

    ws.add_chart(chart, anchor)


def _build_dashboard(wb, cat_data: list, level_data: list) -> None:
    """Build a Dashboard overview sheet with KPIs and charts."""
    ws = wb.create_sheet("Dashboard", 0)  # first sheet
    ws.sheet_properties.tabColor = DARK_BLUE

    # ── Title ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "📊  Цалингийн судалгааны тайлан"
    title_cell.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"Огноо: {datetime.now().strftime('%Y-%m-%d')}"
    sub.font = SUBTITLE_FONT
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # ── KPI cards (row 4–5) ─────────────────────────────────────────────────
    kpi_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    kpi_font_label = Font(name="Calibri", size=9, color=MID_BLUE)
    kpi_font_value = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)

    total_categories = len(cat_data)
    total_levels = len(level_data)
    avg_cat_salary = (sum(r[3] for r in cat_data if r[3]) / total_categories) if total_categories else 0
    total_jobs = sum(r[4] for r in cat_data if r[4])

    kpis = [
        ("Ангилал", str(total_categories)),
        ("Түвшин", str(total_levels)),
        ("Дундаж цалин", f"{avg_cat_salary:,.0f}₮"),
        ("Нийт зар", str(total_jobs)),
    ]

    for i, (label, value) in enumerate(kpis):
        col = i * 2 + 1
        lbl_cell = ws.cell(row=4, column=col, value=label)
        lbl_cell.font = kpi_font_label
        lbl_cell.fill = kpi_fill
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.border = THIN_BORDER
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.cell(row=4, column=col + 1).fill = kpi_fill
        ws.cell(row=4, column=col + 1).border = THIN_BORDER

        val_cell = ws.cell(row=5, column=col, value=value)
        val_cell.font = kpi_font_value
        val_cell.fill = kpi_fill
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = THIN_BORDER
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.cell(row=5, column=col + 1).fill = kpi_fill
        ws.cell(row=5, column=col + 1).border = THIN_BORDER

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 32

    # ── Set column widths up-front ──────────────────────────────────────────
    dash_widths: List[float] = [30, 16, 16, 16, 12]
    _set_column_widths(ws, dash_widths)

    # ── Mini table 1: Top categories by avg salary (row 7+) ────────────────
    ws.cell(row=7, column=1, value="Шилдэг ангилал (дундаж цалингаар)").font = SUBTITLE_FONT
    mini_headers = ["Ангилал", "Мин", "Макс", "Дундаж", "Зарын тоо"]
    for ci, h in enumerate(mini_headers, start=1):
        c = ws.cell(row=8, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGNMENT
        c.border = THIN_BORDER

    sorted_cat = sorted(cat_data, key=lambda x: x[3] or 0, reverse=True)[:8]
    cat_table_start = 9
    for ri, row_data in enumerate(sorted_cat, start=cat_table_start):
        for ci, val in enumerate([row_data[0], row_data[1], row_data[2], row_data[3], row_data[4]], start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.fill = ROW_FILL_EVEN if ri % 2 == 0 else ROW_FILL_ODD
            if ci in (2, 3, 4) and isinstance(val, (int, float)):
                c.number_format = SALARY_FORMAT
                c.alignment = Alignment(horizontal="right")
            elif ci == 5:
                c.alignment = Alignment(horizontal="center")

    cat_table_end = cat_table_start + len(sorted_cat) - 1  # last data row

    # ── Bar chart – references mini-table col D (avg salary) directly ──────
    if sorted_cat:
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Ангиллын дундаж цалин"
        bar.y_axis.numFmt = '#,##0'
        bar.y_axis.title = "₮"
        bar.height = 12
        bar.width = 22
        bar.legend.position = "b"

        # col D = 4 (Дундаж), header at row 8, data rows 9..cat_table_end
        data_ref = Reference(ws, min_col=4, max_col=4, min_row=8, max_row=cat_table_end)
        cats_ref = Reference(ws, min_col=1, min_row=cat_table_start, max_row=cat_table_end)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.series[0].graphicalProperties.solidFill = MID_BLUE

        # Place chart to the right of the table (col G = column 7)
        ws.add_chart(bar, "G7")

    # ── Pie chart – references mini-table col E (job count) directly ───────
    if sorted_cat:
        pie = PieChart()
        pie.style = 10
        pie.title = "Зарын тоо (ангиллаар)"
        pie.height = 12
        pie.width = 18
        if pie.legend is not None:
            pie.legend.position = "r"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = False

        # col E = 5 (Зарын тоо), header at row 8, data rows 9..cat_table_end
        pie_data = Reference(ws, min_col=5, max_col=5, min_row=8, max_row=cat_table_end)
        pie_cats = Reference(ws, min_col=1, min_row=cat_table_start, max_row=cat_table_end)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_cats)

        ws.add_chart(pie, "G27")

    # ── Mini table 2: Levels by avg salary ──────────────────────────────────
    lvl_start_row = max(cat_table_end + 2, 20)  # leave gap
    ws.cell(row=lvl_start_row, column=1, value="Түвшингийн цалин (дундаж цалингаар)").font = SUBTITLE_FONT
    lvl_hdr_row = lvl_start_row + 1
    for ci, h in enumerate(["Түвшин", "Мин", "Макс", "Дундаж", "Зарын тоо"], start=1):
        c = ws.cell(row=lvl_hdr_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGNMENT
        c.border = THIN_BORDER

    sorted_lvl = sorted(level_data, key=lambda x: x[3] or 0, reverse=True)[:8]
    lvl_data_start = lvl_hdr_row + 1
    for ri, row_data in enumerate(sorted_lvl, start=lvl_data_start):
        for ci, val in enumerate([row_data[0], row_data[1], row_data[2], row_data[3], row_data[4]], start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.fill = ROW_FILL_EVEN if ri % 2 == 0 else ROW_FILL_ODD
            if ci in (2, 3, 4) and isinstance(val, (int, float)):
                c.number_format = SALARY_FORMAT
                c.alignment = Alignment(horizontal="right")
            elif ci == 5:
                c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"


# ── Main generator ──────────────────────────────────────────────────────────

def generate_salary_report(output_file: str = "outputs/salary_report.xlsx") -> str:
    engine = _get_engine()
    Session = sessionmaker(bind=engine)
    session = Session()

    try:
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        # ── Load data ──────────────────────────────────────────────────────
        category_rows = session.query(JobCategorySalary).all()
        category_data = [
            [r.job_category, r.min_salary, r.max_salary, r.average_salary,
             r.job_count, r.source_zangia, r.source_lambda, r.month, r.year]
            for r in category_rows
        ]

        level_rows = session.query(JobLevelSalary).all()
        level_data = [
            [r.job_level, r.min_salary, r.max_salary, r.average_salary,
             r.job_count, r.source_zangia, r.source_lambda, r.month, r.year]
            for r in level_rows
        ]

        # ── Dashboard (first sheet) ───────────────────────────────────────
        _build_dashboard(wb, category_data, level_data)

        # ── Category Summary ───────────────────────────────────────────────
        summary_headers = [
            "Ажлын ангилал", "Хамгийн бага цалин", "Хамгийн их цалин",
            "Дундаж цалин", "Зарын тоо", "Zangia", "Lambda Global",
            "Сар", "Жил",
        ]
        summary_widths = [
            WIDTH_PRESETS["name"],
            WIDTH_PRESETS["salary"], WIDTH_PRESETS["salary"], WIDTH_PRESETS["salary"],
            WIDTH_PRESETS["count"], WIDTH_PRESETS["count"], WIDTH_PRESETS["count"],
            WIDTH_PRESETS["period"], WIDTH_PRESETS["period"],
        ]
        n = _write_sheet(
            wb.create_sheet(), "Category Summary",
            summary_headers, category_data,
            col_widths=summary_widths,
            salary_cols=[2, 3, 4],
            number_cols=[5, 6, 7, 8, 9],
        )
        _add_bar_chart(wb["Category Summary"], "Ангиллын цалингийн харьцуулалт",
                       n, 2, 4, anchor="L2")
        _add_pie_chart(wb["Category Summary"], "Зарын тоо (ангиллаар)",
                       n, data_col=5, anchor="L28")

        # ── Category Requirements ──────────────────────────────────────────
        cat_req_rows = session.query(JobCategoryRequirement).all()
        cat_req_data = [[r.job_category, r.requirement_name, r.requirement_details]
                        for r in cat_req_rows]
        _write_sheet(
            wb.create_sheet(), "Category Requirements",
            ["Ажлын ангилал", "Шаардлага", "Дэлгэрэнгүй"],
            cat_req_data,
            col_widths=[WIDTH_PRESETS["name"], WIDTH_PRESETS["text_short"], WIDTH_PRESETS["text_long"]],
        )

        # ── Category Bonuses ───────────────────────────────────────────────
        cat_bonus_rows = session.query(JobCategoryBonus).all()
        cat_bonus_data = [[r.job_category, r.bonus_name, r.bonus_description]
                          for r in cat_bonus_rows]
        _write_sheet(
            wb.create_sheet(), "Category Bonuses",
            ["Ажлын ангилал", "Урамшуулал", "Тайлбар"],
            cat_bonus_data,
            col_widths=[WIDTH_PRESETS["name"], WIDTH_PRESETS["text_short"], WIDTH_PRESETS["text_long"]],
        )

        # ── Level Summary ──────────────────────────────────────────────────
        level_headers = [
            "Албан тушаалын түвшин", "Хамгийн бага цалин", "Хамгийн их цалин",
            "Дундаж цалин", "Зарын тоо", "Zangia", "Lambda Global",
            "Сар", "Жил",
        ]
        n2 = _write_sheet(
            wb.create_sheet(), "Level Summary",
            level_headers, level_data,
            col_widths=summary_widths,
            salary_cols=[2, 3, 4],
            number_cols=[5, 6, 7, 8, 9],
        )
        _add_bar_chart(wb["Level Summary"], "Түвшний цалингийн харьцуулалт",
                       n2, 2, 4, anchor="L2")
        _add_pie_chart(wb["Level Summary"], "Зарын тоо (түвшнээр)",
                       n2, data_col=5, anchor="L28")

        # ── Level Requirements ─────────────────────────────────────────────
        lvl_req_rows = session.query(JobLevelRequirement).all()
        lvl_req_data = [[r.job_level, r.requirement_name, r.requirement_details]
                        for r in lvl_req_rows]
        _write_sheet(
            wb.create_sheet(), "Level Requirements",
            ["Түвшин", "Шаардлага", "Дэлгэрэнгүй"],
            lvl_req_data,
            col_widths=[WIDTH_PRESETS["name"], WIDTH_PRESETS["text_short"], WIDTH_PRESETS["text_long"]],
        )

        # ── Level Bonuses ──────────────────────────────────────────────────
        lvl_bonus_rows = session.query(JobLevelBonus).all()
        lvl_bonus_data = [[r.job_level, r.bonus_name, r.bonus_description]
                          for r in lvl_bonus_rows]
        _write_sheet(
            wb.create_sheet(), "Level Bonuses",
            ["Түвшин", "Урамшуулал", "Тайлбар"],
            lvl_bonus_data,
            col_widths=[WIDTH_PRESETS["name"], WIDTH_PRESETS["text_short"], WIDTH_PRESETS["text_long"]],
        )

        # ── Save ───────────────────────────────────────────────────────────
        os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)
        wb.save(output_file)
        return output_file

    finally:
        session.close()


def main():
    output = generate_salary_report()
    print(f"Report saved to {output}")


if __name__ == "__main__":
    main()
