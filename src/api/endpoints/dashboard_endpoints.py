from __future__ import annotations

import os
import time
from io import BytesIO
from typing import Any, Dict, List, Optional, cast

import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import create_engine, text

from src.api.api_routes import register

_MAIN_CACHE: Optional[pd.DataFrame] = None
_MAIN_CACHE_TS: float = 0.0
_JOBS_CACHE: Optional[pd.DataFrame] = None
_JOBS_CACHE_TS: float = 0.0
_JOBS_FILTER_OPTIONS_CACHE: Optional[Dict[str, List[str]]] = None
_JOBS_FILTER_OPTIONS_CACHE_TS: float = 0.0
_JOBS_COLUMNS_CACHE: Optional[set[str]] = None
_JOBS_COLUMNS_CACHE_TS: float = 0.0


def _engine():
    return create_engine(os.getenv("DATABASE_URI", "sqlite:///products.db"), pool_pre_ping=True)


def _to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip().replace("\u00a0", " ").replace("₮", "").replace(" ", "")
        if not text:
            return None
        if text.count(",") > 1 and "." not in text:
            text = text.replace(",", "")
        elif text.count(".") > 1 and "," not in text:
            text = text.replace(".", "")
        elif "," in text and "." not in text:
            text = text.replace(",", "")
        value = text
    try:
        parsed = float(cast(str | int | float, value))
    except (TypeError, ValueError):
        return None
    if pd.isna(parsed):
        return None
    return parsed


def _json_ready_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    if df.empty:
        return []
    prepared = df.copy().astype(object)
    prepared = prepared.where(pd.notnull(prepared), "")

    datetime_cols = prepared.select_dtypes(include=["datetime64[ns]", "datetimetz"]).columns.tolist()
    for col in datetime_cols:
        prepared[col] = prepared[col].apply(lambda x: x.isoformat() if x else "")

    return cast(List[Dict[str, Any]], prepared.to_dict("records"))


def _load_main_df(refresh: bool = False) -> pd.DataFrame:
    global _MAIN_CACHE, _MAIN_CACHE_TS

    now = time.time()
    if not refresh and _MAIN_CACHE is not None and (now - _MAIN_CACHE_TS) < 45:
        return _MAIN_CACHE.copy()

    df = pd.read_sql("SELECT * FROM salary_calculation_output ORDER BY created_at DESC", _engine())

    for col in [
        "min_salary",
        "max_salary",
        "average_salary",
        "job_count",
        "zangia_count",
        "lambda_count",
        "year",
        "month",
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "created_at" in df.columns:
        df["created_at"] = pd.to_datetime(df["created_at"], errors="coerce")

    if "year" in df.columns and "month" in df.columns:
        df["period"] = df.apply(
            lambda x: f"{int(x['year'])}-{int(x['month']):02d}"
            if pd.notna(x["year"]) and pd.notna(x["month"])
            else None,
            axis=1,
        )

    for col in ["job_count", "zangia_count", "lambda_count"]:
        if col in df.columns:
            df[col] = df[col].fillna(0).astype(int)

    _MAIN_CACHE = df
    _MAIN_CACHE_TS = now
    return df.copy()


def _load_jobs_df(limit: int = 10000, refresh: bool = False) -> pd.DataFrame:
    global _JOBS_CACHE, _JOBS_CACHE_TS

    now = time.time()
    if not refresh and _JOBS_CACHE is not None and (now - _JOBS_CACHE_TS) < 45:
        return _JOBS_CACHE.head(limit).copy()

    existing_cols = _get_jobs_columns(refresh=refresh)
    if not existing_cols:
        return pd.DataFrame()

    select_cols = _jobs_select_columns(existing_cols)
    select_clause = ", ".join(select_cols)

    if "created_at" in existing_cols:
        order_clause = " ORDER BY created_at DESC"
    elif "id" in existing_cols:
        order_clause = " ORDER BY id DESC"
    else:
        order_clause = ""

    query = f"SELECT {select_clause} FROM job_classification_output{order_clause} LIMIT :limit"
    engine = _engine()
    df = pd.read_sql(text(query), engine, params={"limit": int(limit)})

    _JOBS_CACHE = df
    _JOBS_CACHE_TS = now
    return df.copy()


def _first_existing_col(candidates: List[str], existing_cols: set[str]) -> Optional[str]:
    for col in candidates:
        if col in existing_cols:
            return col
    return None


def _get_jobs_columns(refresh: bool = False) -> set[str]:
    global _JOBS_COLUMNS_CACHE, _JOBS_COLUMNS_CACHE_TS

    now = time.time()
    if not refresh and _JOBS_COLUMNS_CACHE is not None and (now - _JOBS_COLUMNS_CACHE_TS) < 600:
        return set(_JOBS_COLUMNS_CACHE)

    engine = _engine()
    cols_df = pd.read_sql("SELECT * FROM job_classification_output LIMIT 1", engine)
    cols = set(cols_df.columns.tolist())
    _JOBS_COLUMNS_CACHE = cols
    _JOBS_COLUMNS_CACHE_TS = now
    return set(cols)


def _jobs_select_columns(existing_cols: set[str]) -> List[str]:
    preferred = [
        "title",
        "source_job",
        "source",
        "platform",
        "site_name",
        "website",
        "company_name",
        "job_level",
        "experience_level",
        "education_level",
        "salary_min",
        "salary_max",
        "requirement_reasoning",
        "requirements",
        "benefits_reasoning",
        "benefits",
        "job_function",
        "job_industry",
        "job_techpack_category",
        "created_at",
        "id",
    ]
    cols = [col for col in preferred if col in existing_cols]
    if cols:
        return cols
    return sorted(existing_cols)


def _distinct_values(engine, column_name: str) -> List[str]:
    query = text(f"SELECT DISTINCT {column_name} AS value FROM job_classification_output WHERE {column_name} IS NOT NULL")
    df = pd.read_sql(query, engine)
    if df.empty or "value" not in df.columns:
        return []
    values = [str(v).strip() for v in df["value"].astype(str).tolist() if str(v).strip()]
    return sorted(list(set(values)))


def _load_jobs_filter_options(refresh: bool = False) -> Dict[str, List[str]]:
    global _JOBS_FILTER_OPTIONS_CACHE, _JOBS_FILTER_OPTIONS_CACHE_TS

    now = time.time()
    if not refresh and _JOBS_FILTER_OPTIONS_CACHE is not None and (now - _JOBS_FILTER_OPTIONS_CACHE_TS) < 300:
        return dict(_JOBS_FILTER_OPTIONS_CACHE)

    engine = _engine()
    existing_cols = _get_jobs_columns(refresh=refresh)
    if not existing_cols:
        return {
            "source": [],
            "function": [],
            "industry": [],
            "level": [],
            "company": [],
        }

    source_col = _first_existing_col(["source_job", "source", "platform", "site_name", "website"], existing_cols)
    function_col = _first_existing_col(["job_function"], existing_cols)
    industry_col = _first_existing_col(["job_industry"], existing_cols)
    level_col = _first_existing_col(["job_level"], existing_cols)
    company_col = _first_existing_col(["company_name"], existing_cols)

    options: Dict[str, List[str]] = {
        "source": _distinct_values(engine, source_col) if source_col else [],
        "function": _distinct_values(engine, function_col) if function_col else [],
        "industry": _distinct_values(engine, industry_col) if industry_col else [],
        "level": _distinct_values(engine, level_col) if level_col else [],
        "company": _distinct_values(engine, company_col) if company_col else [],
    }

    _JOBS_FILTER_OPTIONS_CACHE = options
    _JOBS_FILTER_OPTIONS_CACHE_TS = now
    return dict(options)


def _load_jobs_df_filtered(
    limit: int = 10000,
    refresh: bool = False,
    search: Optional[str] = None,
    source: Optional[str] = None,
    selected_function: Optional[str] = None,
    selected_industry: Optional[str] = None,
    selected_level: Optional[str] = None,
    selected_company: Optional[str] = None,
) -> pd.DataFrame:
    global _JOBS_CACHE, _JOBS_CACHE_TS

    has_filters = any(
        bool(v and str(v).strip())
        for v in [search, source, selected_function, selected_industry, selected_level, selected_company]
    )
    if not has_filters:
        return _load_jobs_df(limit=limit, refresh=refresh)

    engine = _engine()
    existing_cols = _get_jobs_columns(refresh=refresh)
    if not existing_cols:
        return pd.DataFrame()

    source_col = _first_existing_col(["source_job", "source", "platform", "site_name", "website"], existing_cols)
    function_col = _first_existing_col(["job_function"], existing_cols)
    industry_col = _first_existing_col(["job_industry"], existing_cols)
    level_col = _first_existing_col(["job_level"], existing_cols)
    company_col = _first_existing_col(["company_name"], existing_cols)

    where_parts: List[str] = []
    params: Dict[str, str | int] = {}

    def _add_equals(col: Optional[str], value: Optional[str], param_name: str):
        if col and value and str(value).strip():
            where_parts.append(f"LOWER(CAST({col} AS TEXT)) = :{param_name}")
            params[param_name] = str(value).strip().lower()

    _add_equals(source_col, source, "source")
    _add_equals(function_col, selected_function, "job_function")
    _add_equals(industry_col, selected_industry, "job_industry")
    _add_equals(level_col, selected_level, "job_level")
    _add_equals(company_col, selected_company, "company_name")

    if search and str(search).strip():
        q = str(search).strip().lower()
        searchable_cols = [
            col
            for col in [
                "title",
                "source_job",
                "company_name",
                "job_level",
                "experience_level",
                "education_level",
                "requirement_reasoning",
                "requirements",
                "benefits_reasoning",
                "benefits",
                "job_function",
                "job_industry",
                "job_techpack_category",
            ]
            if col in existing_cols
        ]
        if searchable_cols:
            or_parts = [f"LOWER(CAST({col} AS TEXT)) LIKE :search" for col in searchable_cols]
            where_parts.append("(" + " OR ".join(or_parts) + ")")
            params["search"] = f"%{q}%"

    where_clause = f" WHERE {' AND '.join(where_parts)}" if where_parts else ""
    if "created_at" in existing_cols:
        order_clause = " ORDER BY created_at DESC"
    elif "id" in existing_cols:
        order_clause = " ORDER BY id DESC"
    else:
        order_clause = ""

    select_cols = _jobs_select_columns(existing_cols)
    select_clause = ", ".join(select_cols)
    query = f"SELECT {select_clause} FROM job_classification_output{where_clause}{order_clause} LIMIT :limit"
    params["limit"] = int(limit)

    df = pd.read_sql(text(query), engine, params=params)

    _JOBS_CACHE = df
    _JOBS_CACHE_TS = time.time()
    return df.copy()


def _rows_to_excel_bytes(rows: List[Dict[str, Any]], sheet_name: str) -> bytes:
    df = pd.DataFrame(rows)

    for col in ["salary_min", "salary_max", "Доод цалин", "Дээд цалин", "Дундаж цалин"]:
        if col in df.columns:
            df[col] = df[col].apply(_to_float)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)

        money_cols = {"salary_min", "salary_max", "Доод цалин", "Дээд цалин", "Дундаж цалин"}
        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value) in money_cols:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = '#,##0"₮"'
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal="right")

    buffer.seek(0)
    return buffer.read()


@register(name="dashboard/main-data", method="POST", required_keys=[], optional_keys={"refresh": False})
async def dashboard_main_data_handler(data: dict):
    refresh = bool(data.get("refresh", False))
    df = _load_main_df(refresh=refresh)
    return {"items": _json_ready_records(df), "count": int(len(df))}


@register(
    name="dashboard/jobs-data",
    method="POST",
    required_keys=[],
    optional_keys={
        "limit": 10000,
        "refresh": False,
        "search": None,
        "source": None,
        "selected_function": None,
        "selected_industry": None,
        "selected_level": None,
        "selected_company": None,
    },
)
async def dashboard_jobs_data_handler(data: dict):
    limit = int(data.get("limit", 10000) or 10000)
    refresh = bool(data.get("refresh", False))
    df = _load_jobs_df_filtered(
        limit=limit,
        refresh=refresh,
        search=str(data.get("search") or "").strip() or None,
        source=str(data.get("source") or "").strip() or None,
        selected_function=str(data.get("selected_function") or "").strip() or None,
        selected_industry=str(data.get("selected_industry") or "").strip() or None,
        selected_level=str(data.get("selected_level") or "").strip() or None,
        selected_company=str(data.get("selected_company") or "").strip() or None,
    )
    return {"items": _json_ready_records(df.head(limit)), "count": int(len(df))}


@register(name="dashboard/jobs-filter-options", method="POST", required_keys=[], optional_keys={"refresh": False})
async def dashboard_jobs_filter_options_handler(data: dict):
    refresh = bool(data.get("refresh", False))
    options = _load_jobs_filter_options(refresh=refresh)
    return {
        "source_options": [{"label": v, "value": v} for v in options.get("source", [])],
        "function_options": [{"label": v, "value": v} for v in options.get("function", [])],
        "industry_options": [{"label": v, "value": v} for v in options.get("industry", [])],
        "level_options": [{"label": v, "value": v} for v in options.get("level", [])],
        "company_options": [{"label": v, "value": v} for v in options.get("company", [])],
    }


@register(name="download/dashboard-report", method="POST", required_keys=["rows"], optional_keys={"filename": "salary_report.xlsx"})
async def download_dashboard_report_handler(data: dict):
    rows = data.get("rows")
    if not isinstance(rows, list):
        rows = []
    filename = str(data.get("filename") or "salary_report.xlsx")
    excel_bytes = _rows_to_excel_bytes(cast(List[Dict[str, Any]], rows), "Salary_Report")
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@register(name="download/jobs-report", method="POST", required_keys=["rows"], optional_keys={"filename": "jobs_list.xlsx"})
async def download_jobs_report_handler(data: dict):
    rows = data.get("rows")
    if not isinstance(rows, list):
        rows = []
    filename = str(data.get("filename") or "jobs_list.xlsx")
    excel_bytes = _rows_to_excel_bytes(cast(List[Dict[str, Any]], rows), "Jobs_List")
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
