from __future__ import annotations

import json
import os
from datetime import datetime
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from io import BytesIO
from typing import List, Optional

import smtplib
import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import create_engine

from src.dependencies import get_classifier_output_repository
from schemas.base_classifier import JobClassificationOutput, JobRequirement, JobBenefit

load_dotenv()

app = FastAPI(title="Salary Agent Email API", version="1.0.0")


class EmailRequest(BaseModel):
    to_email: str = Field(..., description="Recipient email address")
    subject: Optional[str] = Field(None, description="Email subject")

class SalaryReportRequest(BaseModel):
    to_email: str = Field(..., description="Recipient email address")
    subject: Optional[str] = Field(None, description="Email subject")
    type: Optional[str] = Field("function", description="Type filter: function, job_level, industry, or techpack_category")


def _format_requirements(requirements: List[object]) -> str:
    if not requirements:
        return ""
    formatted: List[str] = []
    for req in requirements:
        if isinstance(req, JobRequirement):
            importance = f" ({req.importance})" if req.importance else ""
            formatted.append(f"{req.name}: {req.details}{importance}")
        elif isinstance(req, dict):
            name = str(req.get("name", "")).strip()
            details = str(req.get("details", "")).strip()
            importance = str(req.get("importance", "")).strip()
            suffix = f" ({importance})" if importance else ""
            text = ": ".join(part for part in [name, details] if part)
            formatted.append(f"{text}{suffix}" if text else "")
        else:
            formatted.append(str(req))
    return ", ".join([item for item in formatted if item])


def _format_benefits(benefits: List[object]) -> str:
    if not benefits:
        return ""
    formatted: List[str] = []
    for benefit in benefits:
        if isinstance(benefit, JobBenefit):
            value = f" (Үнэлгээ: {benefit.monetary_value} MNT)" if benefit.monetary_value else ""
            formatted.append(f"{benefit.name}: {benefit.description}{value}")
        elif isinstance(benefit, dict):
            name = str(benefit.get("name", "")).strip()
            description = str(benefit.get("description", "")).strip()
            value = benefit.get("monetary_value")
            suffix = f" (Үнэлгээ: {value} MNT)" if value else ""
            text = ": ".join(part for part in [name, description] if part)
            formatted.append(f"{text}{suffix}" if text else "")
        else:
            formatted.append(str(benefit))
    return ", ".join([item for item in formatted if item])


def _enum_value(value: Optional[object]) -> str:
    if value is None:
        return ""
    return str(getattr(value, "value", value))


def _coerce_list(value: object) -> List[object]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    return parsed
            except json.JSONDecodeError:
                return [value]
        return [value]
    return [value]

def _build_excel() -> bytes:
    classifications = get_classifier_output_repository().get_all()

    workbook = Workbook()
    sheet: Worksheet = workbook.active  # type: ignore
    sheet.title = "Job Classifications" 

    headers = [
        "Title",
        "Job Function",
        "Job Industry",
        "Job Level",
        "Job Techpack Category",
        "Experience Level",
        "Education Level",
        "Salary Min",
        "Salary Max",
        "Company",
        "Requirements",
        "Benefits",
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    body_alignment = Alignment(vertical="top", wrap_text=True)

    for item in classifications:
        # Handle requirements formatting
        formatted_requirements = _format_requirements(_coerce_list(item.requirements))
        
        # Handle benefits formatting
        formatted_benefits = _format_benefits(_coerce_list(item.benefits))
        
        sheet.append(
            [
                item.title,
                _enum_value(item.job_function),
                _enum_value(item.job_industry),
                _enum_value(item.job_level),
                _enum_value(item.job_techpack_category),
                _enum_value(item.experience_level),
                _enum_value(item.education_level),
                item.salary_min,
                item.salary_max,
                item.company_name or "",
                formatted_requirements,
                formatted_benefits,
            ]
        )

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = body_alignment

    column_widths = {
        1: 28,
        2: 20,
        3: 24,
        4: 18,
        5: 28,
        6: 18,
        7: 18,
        8: 14,
        9: 14,
        10: 20,
        11: 40,
        12: 40,
    }

    for col_idx, width in column_widths.items():
        sheet.column_dimensions[chr(64 + col_idx)].width = width

    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()


def _build_salary_excel(type_filter: str = "function", title_filter: Optional[str] = None) -> bytes:
    """Build Excel report for salary calculation data."""
    engine = create_engine(
        os.getenv("DATABASE_URI", "sqlite:///products.db"),
        pool_pre_ping=True,
    )
    query = "SELECT * FROM salary_calculation_output ORDER BY created_at DESC"
    df = pd.read_sql(query, engine)
    
    # Apply filters
    if "type" in df.columns:
        df = df[df["type"] == type_filter]
    
    if title_filter and "title" in df.columns:
        df = df[df["title"] == title_filter]
    
    # Exclude "All" titles
    if "title" in df.columns:
        df = df[~df["title"].astype(str).str.startswith("All ")]
    
    # Get latest per title
    if "created_at" in df.columns and df["created_at"].notna().any():
        df = df.sort_values("created_at").drop_duplicates(subset=["title"], keep="last")
    
    # Prepare display columns
    display_df = df.rename(
        columns={
            "title": "Ажлын ангилал",
            "min_salary": "Доод цалин",
            "max_salary": "Дээд цалин",
            "average_salary": "Дундаж цалин",
            "job_count": "Зарын тоо",
            "zangia_count": "Zangia",
            "lambda_count": "Lambda",
        }
    )
    
    columns_to_export = [
        "Ажлын ангилал",
        "Доод цалин",
        "Дээд цалин",
        "Дундаж цалин",
        "Зарын тоо",
        "Zangia",
        "Lambda",
    ]
    display_df = display_df[[col for col in columns_to_export if col in display_df.columns]]
    
    # Create Excel file
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        display_df.to_excel(writer, index=False, sheet_name="Salary_Report")
        ws = writer.sheets["Salary_Report"]
        
        # Style header
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)
        
        # Format money columns
        money_cols = {"Доод цалин", "Дээд цалин", "Дундаж цалин"}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value in money_cols:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = '#,##0"₮"'
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal="right")
            if cell.value in {"Зарын тоо", "Zangia", "Lambda"}:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal="center")
    
    buffer.seek(0)
    return buffer.read()


def _send_email(to_email: str, subject: str, attachment: bytes, filename: str) -> None:
    sender_email = os.getenv("SENDER_EMAIL", "itgel6708@gmail.com")
    app_password = os.getenv("GMAIL_APP_PASSWORD", "")

    # Allow overriding recipients via env vars, but keep current defaults.
    receiver_emails = [to_email]
    # cc_emails = _parse_email_list(os.getenv("CC_EMAILS") or os.getenv("CC_EMAIL"), fallback=[])
     # Root container must be 'mixed' when there are attachments.
    msg = MIMEMultipart("mixed")
    msg["From"] = sender_email
    msg["To"] = ", ".join(receiver_emails)
    msg["Subject"] = subject

    # Attach the Excel file.
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(attachment)    
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)
    # Send the email via SMTP server.
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender_email, app_password)
        server.send_message(msg)

    


@app.get("/health")
async def health_check() -> dict:
    return {"status": "ok"}


@app.post("/email/job-classifications")
async def email_job_classifications(request: EmailRequest) -> dict:


    subject = request.subject or "Job Classification Results"

    try:
        attachment = _build_excel()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"job_classifications_{timestamp}.xlsx"
        _send_email(request.to_email, subject, attachment, filename)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return {"status": "sent", "to": request.to_email, "filename": filename}

@app.post("/download/salary-report")
async def download_salary_report_post(request: SalaryReportRequest):
    """Download salary report as Excel file (POST version).
    
    Args:
        request: Request with type and optional title filter
    
    Returns:
        Excel file with salary report data
    """
    try:
        excel_bytes = _build_salary_excel(type_filter=request.type or "function")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"salary_report_{timestamp}.xlsx"
        
        return StreamingResponse(
            BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/email/salary-report")
async def email_salary_report(request: SalaryReportRequest) -> dict:
    """Generate salary report and send it via email.
    
    Args:
        request: Request with recipient email, subject, type, and optional title filter
    Returns:
        Status of email sending
    """

    subject = request.subject or "Salary Report"

    

    try:
        excel_bytes = _build_salary_excel(type_filter=request.type or "function")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"salary_report_{timestamp}.xlsx"
        _send_email(request.to_email, subject, excel_bytes, filename)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return {"status": "sent", "to": request.to_email, "filename": filename}
